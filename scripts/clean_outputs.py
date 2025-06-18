# scripts/clean_outputs.py

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from deduper import dedupe_contacts

OUTPUT_FOLDER = "Output"
CLEANED_FOLDER = "Cleaned"
LOG_PATH = "logs/flagged_for_review.csv"

def drop_struck_duplicates(df):
    df = df.copy()
    df["Normalized Name"] = (df["First Name"].fillna('') + " " + df["Last Name"].fillna('')).str.lower().str.strip()
    seen = set()
    keep_rows = []

    for _, row in df.iterrows():
        name = row["Normalized Name"]
        strike = row.get("Strike", False)
        if strike:
            continue
        if name not in seen:
            keep_rows.append(row)
            seen.add(name)
    return pd.DataFrame(keep_rows)

def format_cleaned_output(df_cleaned, df_original, firm_name, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    font_name = "Aptos Narrow"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True, name=font_name, size=11)
    italic_font = Font(italic=True, name=font_name, size=11)
    strike_font = Font(strike=True, name=font_name, size=11)
    regular_font = Font(name=font_name, size=11)
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    thin_border = Border(bottom=Side(style="thin"))
    double_border = Border(bottom=Side(style="double"))

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=2)
    title_cell.value = firm_name
    title_cell.font = italic_font
    title_cell.alignment = center_align

    headers = ["First Name", "Last Name", "Title", "Account Name", "Notes"]
    for col_num, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = left_align

    df_original["Full Name"] = df_original["First Name"].fillna('') + " " + df_original["Last Name"].fillna('')
    df_original["Normalized Name"] = df_original["Full Name"].str.lower().str.strip()

    main_names = df_original[
        (df_original.get("Highlight", False) == False) &
        (df_original.get("Strike", False) == False)
    ]["Normalized Name"].tolist()

    new_names = df_original[df_original.get("Highlight", False)]["Normalized Name"].tolist()
    removed_df = df_original[df_original.get("Strike", False)].copy()
    removed_df = removed_df[~removed_df["Normalized Name"].isin(main_names)]

    row_ptr = 4
    for _, row in df_cleaned.iterrows():
        full_name = f"{str(row['First Name']).strip()} {str(row['Last Name']).strip()}".lower()
        if full_name in removed_df["Normalized Name"].tolist():
            continue

        is_new = full_name in new_names

        values = [
            row["First Name"],
            row["Last Name"],
            row["Title"],
            firm_name,
            row.get("Notes", "")
        ]

        for col, val in enumerate(values, start=2):
            cell = ws.cell(row=row_ptr, column=col)
            cell.value = val
            cell.font = regular_font
            cell.alignment = left_align
            if is_new:
                cell.fill = yellow_fill
        row_ptr += 1

    row_ptr += 1
    ws.cell(row=row_ptr, column=2).value = "Not Listed"
    ws.cell(row=row_ptr, column=2).font = Font(bold=True, name=font_name, size=11)
    ws.cell(row=row_ptr, column=2).border = double_border
    row_ptr += 1

    for _, row in removed_df.iterrows():
        values = [
            row["First Name"],
            row["Last Name"],
            row["Title"],
            firm_name,
            row.get("Notes", "")
        ]
        for col, val in enumerate(values, start=2):
            cell = ws.cell(row=row_ptr, column=col)
            cell.value = val
            cell.font = strike_font
            cell.alignment = left_align
        row_ptr += 1

    wb.save(output_path)

def clean_all_outputs(selected_firms=None):
    os.makedirs(CLEANED_FOLDER, exist_ok=True)
    files = [f for f in os.listdir(OUTPUT_FOLDER) if f.endswith("_updated_contacts.xlsx") and not f.startswith("~$")]

    if selected_firms:
        files = [f for f in files if f.replace("_updated_contacts.xlsx", "") in selected_firms]

    for file in files:
        firm_name = file.replace("_updated_contacts.xlsx", "")
        input_path = os.path.join(OUTPUT_FOLDER, file)
        output_path = os.path.join(CLEANED_FOLDER, f"{firm_name}.xlsx")

        try:
            df_original = pd.read_excel(input_path).fillna('')
            df_pre_deduped = drop_struck_duplicates(df_original.copy())
            df_cleaned = df_pre_deduped.drop(columns=["Highlight", "Strike", "Normalized Name", "Full Name"], errors="ignore")
            df_cleaned = dedupe_contacts(df_cleaned, firm_name)
            df_cleaned["Account Name"] = firm_name
            format_cleaned_output(df_cleaned, df_original, firm_name, output_path)
            print(f"Cleaned & formatted: {firm_name}")
        except Exception as e:
            print(f"{firm_name} — ERROR: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        clean_all_outputs(selected_firms=sys.argv[1:])
    else:
        clean_all_outputs()
