import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def write_output(data_list, file_path):
    today_str = datetime.today().strftime("%Y-%m-%d")

    processed_data = []
    for row in data_list:
        try:
            # Safe full name logic with debug
            first = str(row.get("First Name", "") or "").strip()
            last = str(row.get("Last Name", "") or "").strip()
            full_name = f"{first} {last}"
            names = full_name.strip().split()
            first_name = names[0] if len(names) > 0 else ""
            last_name = " ".join(names[1:]) if len(names) > 1 else ""
        except Exception as e:
            print("❌ ERROR in write_output() while building full name:")
            print("Row that caused it:", row)
            print("Error message:", e)
            raise

        is_new = row.get("Highlight", False)
        is_strike = row.get("Strike", False)

        processed_data.append({
            "First Name": first_name,
            "Last Name": last_name,
            "Title": row.get("Title", ""),
            "Account Name": row.get("Account Name", ""),
            "Highlight": is_new,
            "Strike": is_strike,
            "Notes": f"New as of {today_str}" if is_new else row.get("Notes", "")
        })

    # Sort to put new contacts at the bottom
    processed_data.sort(key=lambda x: x["Highlight"])

    df = pd.DataFrame(processed_data)
    df.to_excel(file_path, index=False)

    # Apply Excel formatting
    wb = load_workbook(file_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for idx, row in enumerate(processed_data, start=2):  # Start after header
        is_new = row.get("Highlight", False)
        is_strike = row.get("Strike", False)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=idx, column=col)
            if is_new:
                cell.fill = yellow_fill
            if is_strike:
                cell.font = Font(strike=True)

    wb.save(file_path)
