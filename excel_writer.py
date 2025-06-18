# excel_writer.py

import openpyxl
from openpyxl.styles import PatternFill
import os

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def write_team_excel(firm_name, nvca_contacts, new_contacts, output_dir):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Team Contacts"

    # Add header
    ws.append(["Name", "Title"])

    # Add existing NVCA contacts
    for row in nvca_contacts:
        ws.append([row["name"], row["title"]])

    # Add new team members with yellow highlight
    for row in new_contacts:
        ws.append([row["name"], row["title"]])
        for cell in ws[ws.max_row]:
            cell.fill = YELLOW_FILL

    # Save to path
    safe_name = firm_name.replace("/", "_").replace("\\", "_")
    output_path = os.path.join(output_dir, f"{safe_name}_Updated.xlsx")
    wb.save(output_path)
